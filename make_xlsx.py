"""
Convert results.csv to results.xlsx with per-row coloring:
  - Green row  : S wins (fewer nodes than S_ab)
  - Red row    : S_ab wins
  - Light gray : tie
  - Yellow tint on the better cell in each (ab, S) metric column

Also adds a Summary sheet with per-batch aggregates.
"""
import csv, re, sys
from collections import defaultdict
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

IN_CSV = "/home/claude/tableau/results.csv"
OUT_XLSX = "/mnt/user-data/outputs/results.xlsx"

# Colors
GREEN_ROW   = "DDF7DD"   # S wins
RED_ROW     = "FBE2DD"   # S_ab wins
GRAY_ROW    = "EDEDED"   # tie
HDR_FILL    = "305496"   # dark blue
HDR_FONT    = "FFFFFF"   # white
HL_BETTER   = "FFEB9C"   # yellow tint for the winning cell within a metric pair

BORDER = Border(
    left=Side(style="thin", color="CCCCCC"),
    right=Side(style="thin", color="CCCCCC"),
    top=Side(style="thin", color="CCCCCC"),
    bottom=Side(style="thin", color="CCCCCC"),
)

rows = list(csv.DictReader(open(IN_CSV)))
print(f"Read {len(rows)} rows from {IN_CSV}")

wb = Workbook()

# -----------------------------------------------------------------------------
# Sheet 1: Results (per-problem)
# -----------------------------------------------------------------------------
ws = wb.active
ws.title = "Results"

# Column order: problem, winner badge, then ab/S metric pairs side-by-side.
COLUMNS = [
    ("problem",            "Problem",        "txt"),
    ("winner",             "Winner",         "txt"),
    ("ab_nodes",           "S_ab nodes",     "num"),
    ("s_nodes",            "S nodes",        "num"),
    ("ab_expansions",      "S_ab exp",       "num"),
    ("s_expansions",       "S exp",          "num"),
    ("ab_betas",           "S_ab betas",     "num"),
    ("s_betas",            "S betas",        "num"),
    ("ab_depth_max",       "S_ab depth",     "num"),
    ("s_depth_max",        "S depth",        "num"),
    ("ab_branches_closed", "S_ab br.cl",     "num"),
    ("s_branches_closed",  "S br.cl",        "num"),
    ("ab_time_ms",         "S_ab ms",        "f"),
    ("s_time_ms",          "S ms",           "f"),
    ("ab_closed",          "S_ab closed",    "bool"),
    ("s_closed",           "S closed",       "bool"),
]

# Header row
for col_idx, (_, header, _) in enumerate(COLUMNS, start=1):
    c = ws.cell(row=1, column=col_idx, value=header)
    c.font = Font(name="Arial", bold=True, color=HDR_FONT)
    c.fill = PatternFill("solid", start_color=HDR_FILL)
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.border = BORDER

# Pairs of columns where smaller-is-better; key=metric name, value=(ab_col, s_col)
METRIC_PAIRS = [
    ("nodes",      "ab_nodes",      "s_nodes"),
    ("exp",        "ab_expansions", "s_expansions"),
    ("betas",      "ab_betas",      "s_betas"),
    ("depth",      "ab_depth_max",  "s_depth_max"),
    ("ms",         "ab_time_ms",    "s_time_ms"),
]

# Column index lookup
col_index = {key: i + 1 for i, (key, _, _) in enumerate(COLUMNS)}

def row_winner(r):
    """Return 'S', 'S_ab', or 'tie' based on node count (the primary metric)."""
    s, ab = int(r["s_nodes"]), int(r["ab_nodes"])
    if s < ab:  return "S"
    if ab < s:  return "S_ab"
    return "tie"

# Sort same as bench.py
def sort_key(r):
    m = re.match(r"pel(\d+)(?:_b(\d+))?$", r["problem"])
    if not m: return (999, 999, r["problem"])
    return (int(m.group(2)) if m.group(2) else 0, int(m.group(1)), r["problem"])

rows.sort(key=sort_key)

for excel_row, r in enumerate(rows, start=2):
    winner = row_winner(r)
    if winner == "S":
        row_fill = PatternFill("solid", start_color=GREEN_ROW)
    elif winner == "S_ab":
        row_fill = PatternFill("solid", start_color=RED_ROW)
    else:
        row_fill = PatternFill("solid", start_color=GRAY_ROW)

    for col_idx, (key, _, kind) in enumerate(COLUMNS, start=1):
        if key == "winner":
            val = winner
        else:
            raw = r[key]
            if kind == "num":
                val = int(raw)
            elif kind == "f":
                val = float(raw)
            elif kind == "bool":
                val = "yes" if raw == "1" else "no"
            else:
                val = raw
        c = ws.cell(row=excel_row, column=col_idx, value=val)
        c.fill = row_fill
        c.alignment = Alignment(horizontal="center" if kind != "txt" else "left")
        c.border = BORDER
        c.font = Font(name="Arial")
        if kind == "f":
            c.number_format = "0.000"

    # Within each metric pair, highlight the better (smaller) cell in yellow.
    for _, ab_key, s_key in METRIC_PAIRS:
        ab_val = float(r[ab_key])
        s_val  = float(r[s_key])
        ab_cell = ws.cell(row=excel_row, column=col_index[ab_key])
        s_cell  = ws.cell(row=excel_row, column=col_index[s_key])
        if s_val < ab_val:
            s_cell.fill = PatternFill("solid", start_color=HL_BETTER)
            s_cell.font = Font(name="Arial", bold=True)
        elif ab_val < s_val:
            ab_cell.fill = PatternFill("solid", start_color=HL_BETTER)
            ab_cell.font = Font(name="Arial", bold=True)

    # Bold the winner badge
    winner_cell = ws.cell(row=excel_row, column=col_index["winner"])
    winner_cell.font = Font(name="Arial", bold=True)

# Column widths
widths = {
    "problem": 14, "winner": 9,
}
default_w = 11
for col_idx, (key, header, _) in enumerate(COLUMNS, start=1):
    ws.column_dimensions[get_column_letter(col_idx)].width = widths.get(key, default_w)

ws.freeze_panes = "C2"   # freeze first row + problem & winner columns

# -----------------------------------------------------------------------------
# Sheet 2: Summary (per-batch + overall)
# -----------------------------------------------------------------------------
ws2 = wb.create_sheet("Summary")

batch_stats = defaultdict(lambda: {"n": 0, "ab_nodes": 0, "s_nodes": 0,
                                   "ab_exp": 0, "s_exp": 0,
                                   "ab_ms": 0.0, "s_ms": 0.0,
                                   "wins_s": 0, "wins_ab": 0, "ties": 0})

for r in rows:
    m = re.match(r"pel\d+(?:_b(\d+))?$", r["problem"])
    b = m.group(1) if m and m.group(1) else "0"
    bs = batch_stats[b]
    bs["n"] += 1
    bs["ab_nodes"] += int(r["ab_nodes"]); bs["s_nodes"] += int(r["s_nodes"])
    bs["ab_exp"]   += int(r["ab_expansions"]); bs["s_exp"] += int(r["s_expansions"])
    bs["ab_ms"]    += float(r["ab_time_ms"]); bs["s_ms"]   += float(r["s_time_ms"])
    w = row_winner(r)
    if w == "S":     bs["wins_s"]  += 1
    elif w == "S_ab":bs["wins_ab"] += 1
    else:            bs["ties"]    += 1

SUMMARY_COLS = ["Batch", "N",
                "S_ab nodes", "S nodes", "Ratio S/S_ab",
                "S_ab exp", "S exp",
                "S_ab ms", "S ms",
                "S wins", "S_ab wins", "Ties"]

for i, h in enumerate(SUMMARY_COLS, start=1):
    c = ws2.cell(row=1, column=i, value=h)
    c.font = Font(name="Arial", bold=True, color=HDR_FONT)
    c.fill = PatternFill("solid", start_color=HDR_FILL)
    c.alignment = Alignment(horizontal="center")
    c.border = BORDER

def write_summary_row(r, label, bs):
    ratio = bs["s_nodes"] / bs["ab_nodes"] if bs["ab_nodes"] else 0
    values = [label, bs["n"],
              bs["ab_nodes"], bs["s_nodes"], ratio,
              bs["ab_exp"], bs["s_exp"],
              bs["ab_ms"], bs["s_ms"],
              bs["wins_s"], bs["wins_ab"], bs["ties"]]
    # Color row according to which heuristic wins on total nodes
    if ratio < 1.0:
        row_fill = PatternFill("solid", start_color=GREEN_ROW)
    elif ratio > 1.0:
        row_fill = PatternFill("solid", start_color=RED_ROW)
    else:
        row_fill = PatternFill("solid", start_color=GRAY_ROW)
    for i, v in enumerate(values, start=1):
        c = ws2.cell(row=r, column=i, value=v)
        c.fill = row_fill
        c.border = BORDER
        c.font = Font(name="Arial", bold=(i==1))
        c.alignment = Alignment(horizontal="left" if i==1 else "center")
        if i == 5:                # ratio column
            c.number_format = "0.000"
        elif i in (8, 9):         # ms columns
            c.number_format = "0.0"
    # Highlight winning side of the totals
    ab_nodes_cell = ws2.cell(row=r, column=3)
    s_nodes_cell  = ws2.cell(row=r, column=4)
    if bs["s_nodes"] < bs["ab_nodes"]:
        s_nodes_cell.fill = PatternFill("solid", start_color=HL_BETTER)
        s_nodes_cell.font = Font(name="Arial", bold=True)
    elif bs["ab_nodes"] < bs["s_nodes"]:
        ab_nodes_cell.fill = PatternFill("solid", start_color=HL_BETTER)
        ab_nodes_cell.font = Font(name="Arial", bold=True)

excel_row = 2
LABELS = {
    "0": "Originals",
    "1": "B1 (asymmetric depth)",
    "2": "B2 (mixed asymmetric)",
    "3": "B3 (strict chain)",
    "4": "B4 (one huge P)",
    "5": "B5 (recursive containment)",
}
for b in sorted(batch_stats.keys(), key=lambda x: int(x)):
    label = LABELS.get(b, f"Batch {b}")
    write_summary_row(excel_row, label, batch_stats[b])
    excel_row += 1

# Overall totals
overall = {"n": 0, "ab_nodes": 0, "s_nodes": 0, "ab_exp": 0, "s_exp": 0,
           "ab_ms": 0.0, "s_ms": 0.0, "wins_s": 0, "wins_ab": 0, "ties": 0}
for bs in batch_stats.values():
    for k in overall:
        overall[k] += bs[k]
write_summary_row(excel_row, "ALL", overall)

# Add a small legend
excel_row += 3
ws2.cell(row=excel_row, column=1, value="Legend").font = Font(name="Arial", bold=True, size=12)
excel_row += 1
legend = [
    ("S wins (fewer nodes than S_ab)", GREEN_ROW),
    ("S_ab wins", RED_ROW),
    ("Tie",       GRAY_ROW),
    ("Better cell within a metric pair", HL_BETTER),
]
for label, color in legend:
    c = ws2.cell(row=excel_row, column=1, value=label)
    c.fill = PatternFill("solid", start_color=color)
    c.border = BORDER
    c.font = Font(name="Arial")
    excel_row += 1

# Column widths for summary
summary_widths = [12, 5, 12, 11, 14, 11, 11, 11, 11, 9, 11, 7]
for i, w in enumerate(summary_widths, start=1):
    ws2.column_dimensions[get_column_letter(i)].width = w

ws2.freeze_panes = "A2"

# -----------------------------------------------------------------------------
wb.save(OUT_XLSX)
print(f"Wrote {OUT_XLSX}")
